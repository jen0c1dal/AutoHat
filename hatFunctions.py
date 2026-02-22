"""Backend logic for AutoHat app"""

from __future__ import annotations

# Built in libraries
import datetime as dt
from dataclasses import dataclass, field
from enum import Enum
import math
import os
import random as rd

# Third party libraries
import pandas as pd
from openpyxl import Workbook, load_workbook


class Gender(Enum):
    """Enum for player gender"""

    MALE = 'male'
    FEMALE = 'female'


class SkillLevel(Enum):
    """Class representing a skill level with level and text"""

    def __init__(self, level: int, text: str):
        """Initialize a skill level"""

        self.level = level
        self.text = text


class Throws(SkillLevel):
    """Enum for throwing skill levels"""

    NOOB = (1, "I've thrown a frisbee before.")
    COMPETENT = (2, "I can throw a forehand and backhand, even if they're occasionally wobbly.")
    PRO = (3, "Accurate with standard throws; I know what IO and OI mean.")
    SCOOBER_GOD = (4, "All the throws; I will destroy you with my full-field scoobers.")


class Experience(SkillLevel):
    """Enum for experience skill levels"""

    ROOKIE = (1, "Rookie")
    PICKUP = (2, "Pickup player")
    CLUB = (3, "Club (Sectionals) / Masters (Nationals) / High School (State / Nationals)")
    PRO = (4, "Club player (Regionals / Nationals)")


class Endurance(SkillLevel):
    """Enum for endurance skill levels"""

    BACKUP = (1, "I like to rest for a few points in between the points that I play.")
    LINE = (2, "I can play about every other point at full speed.")
    ANCHOR = (3, "I can play a few points in a row at full speed before I need a rest.")
    SAVAGE = (4, "I actually prefer to play savage.")


class Athletics(SkillLevel):
    """Enum for athleticism skill levels"""

    UNFIT = (1, "Out of shape; mostly I'm here to heckle.")
    FIT = (2, "Somewhat athletic; I can usually get open when I make a cut.")
    FAST = (3, "Quite athletic; I have no difficulty getting open when I make cuts.")
    ELITE = (4, "Very athletic; my two settings are Sprint and Horizontal.")


def skill_match(text: str, enum_type) -> Enum:
    """Match text to enum value"""
    for enum in enum_type:
        if enum.text == text:
            return enum.level

    return 0


@dataclass
class Player:
    """Class representing a player"""

    name: str
    gender: Gender
    rank: int

    def __lt__(self, other) -> bool:
        """Compare players by rank"""

        return self.rank < other.rank

    def to_dict(self) -> dict:
        """Convert player to dictionary"""

        return {
            'name': self.name,
            'gender': self.gender.value,
            'rank': self.rank,
        }


@dataclass
class Roster:
    players: list[Player]

    def sort_by_name(self):
        self.players.sort(key=lambda p: p.name)

    def get_player_by_name(self, name: str) -> Player | None:
        return next((p for p in self.players if p.name == name), None)


@dataclass
class PlayerGroup:
    """Class to baggage multiple players together"""

    players: list[Player] = field(default_factory=list)

    def __post_init__(self):
        if not self.players:
            raise ValueError('PlayerGroup must have at least one player')

    @staticmethod
    def calc_mean_rank(roster: list[Player]) -> float:
        """Calculate mean rank of players"""
        return sum(p.rank for p in roster) / len(roster)

    @property
    def mean_rank(self) -> float:
        """Get the mean rank of the players in the group"""
        return self.calc_mean_rank(self.players)

    @property
    def num_fmp(self) -> int:
        """Get the number of female players on the team"""
        return len([p for p in self.players if p.gender == Gender.FEMALE])

    @property
    def num_mmp(self) -> int:
        """Get the number of male players on the team"""
        return len([p for p in self.players if p.gender == Gender.MALE])

    @property
    def num_players(self) -> int:
        return len(self.players)

    def __init__(self, players: list[Player] | list[str], roster: Roster | None = None, **data):
        if isinstance(players[0], str):
            # From list[str] and roster
            self.players = PlayerGroup.create_group(players, roster)
        else:
            # From list[Player]
            self.players = players
        self.__post_init__()

    @staticmethod
    def create_group(players: list[str], roster: Roster):
        """Create a group of players from names and roster"""
        group = []
        for p in players:
            player = roster.get_player_by_name(p)
            if player:
                group.append(player)
        return group


    def add_players(self, players: Player | list[Player]):
        """Add players to the group"""
        if isinstance(players, Player):
            self.players.append(players)
        else:
            self.players.extend(players)

    def remove_player(self, player: Player):
        """Remove a player from the group"""
        if player in self.players:
            self.players.remove(player)

    def __lt__(self, other) -> bool:
        """Compare groups by mean rank"""

        return self.mean_rank < other.mean_rank


class Team(PlayerGroup):
    """Class representing a finalized team with players"""

    def __init__(self, players: list[Player], **data):
        # Initialize as PlayerGroup but skip the string processing
        super().__init__(players, None, **data)

    def __str__(self) -> str:
        """String representation of the team"""
        return f"Team: {len(self.players)} players, Avg Rank: {self.mean_rank:.1f}"


def import_roster(filepath: str) -> Roster:
    """Import roster from CSV file"""
    df = pd.read_csv(filepath)
    players = []
    for _, row in df.iterrows():
        throws = skill_match(row['throws'], Throws)
        experience = skill_match(row['experience'], Experience)
        endurance = skill_match(row['endurance'], Endurance)
        athleticism = skill_match(row['athleticism'], Athletics)
        name = f"{row['first_name']} {row['last_name']}"
        rank = throws + experience + endurance + athleticism
        players.append(Player(name=name, gender=Gender(row['gender']), rank=rank))
    return Roster(players)


def launch_checkin(data_in_path):
    """Launch check-in by importing and sorting roster"""
    roster = import_roster(data_in_path)
    roster.sort_by_name()
    return roster


def assign_players(mean_rank: float, roster: list[Player], teams: list[PlayerGroup], num_teams: int, team_index: int = 0) -> int:
    """Assign players to teams based on rank"""
    while len(roster) > 0:
        if teams[team_index].mean_rank > mean_rank:
            player = pop_random_player(roster, math.ceil(len(roster) / 2), len(roster) - 1)
        else:
            player = pop_random_player(roster, 0, math.floor(len(roster) / 2))
        teams[team_index].add_players(player)
        team_index = (team_index + 1) % num_teams

    return team_index


def pop_random_player(roster: list[Player], begin: int, end: int) -> Player:
    """Pop a random player from roster within range"""
    if len(roster) == 1:
        return roster.pop(0)
    return roster.pop(rd.randint(begin, end))


def add_baggages_to_teams(teams: list[PlayerGroup], baggages: list[PlayerGroup], mean_rank:float):
    """Add baggages to teams"""
    while len(baggages) > 0:
        for t in teams:
            if len(baggages) > 0:
                baggage = baggages.pop() if t.mean_rank > mean_rank else baggages.pop(0)
                t.add_players(baggage.players)


def balance_attribute(teams: list[PlayerGroup], roster: list[Player], mean_rank: float, get_attr):
    """Helper function to balance teams based on a given attribute"""
    max_attr = max(get_attr(t) for t in teams)
    min_attr = min(get_attr(t) for t in teams)
    while min_attr < max_attr and len(roster) > 0:
        for t in teams:
            if get_attr(t) < max_attr:
                if t.mean_rank > mean_rank:
                    t.add_players(pop_random_player(roster, math.ceil(len(roster) / 2), len(roster) - 1))
                else:
                    t.add_players(pop_random_player(roster, 0, math.floor(len(roster) / 2)))
            min_attr = min(get_attr(t) for t in teams)


def balance_teams(teams: list[PlayerGroup], m_roster: list[Player], f_roster: list[Player], mean_rank: float):
    """Balance teams by gender and player count"""
    balance_attribute(teams, f_roster, mean_rank, lambda t: t.num_fmp)
    balance_attribute(teams, m_roster, mean_rank, lambda t: t.num_players)


def add_drop_in(name: str, gender: str, rank: str) -> Player:
    """Add drop-in player to roster"""
    return Player(name=name.title(), gender=Gender(gender), rank=int(rank))


def create_baggage(players: list[str], roster: Roster) -> PlayerGroup:
    return PlayerGroup(players, roster)


def generate_teams(players: list[Player], save_directory: str, num_teams: int, baggages: list[PlayerGroup]):
    """Main function to generate a given number of teams teams from the list of checked in players"""
    mean_rank = PlayerGroup.calc_mean_rank(players)

    # Split the roster into rosters of men and women
    men = [p for p in players if p.gender == Gender.MALE]
    women = [p for p in players if p.gender == Gender.FEMALE]

    men.sort(reverse=True)
    women.sort(reverse=True)

    if len(men) >= num_teams:
        # Assign top players to teams
        teams = [PlayerGroup([men.pop(0)]) for _ in range(num_teams)]
        # Add random players
        for i in range(num_teams):
            teams[i].add_players(pop_random_player(men, 0, len(men) - 1))
    else:
        # Assign top women to teams
        teams = [PlayerGroup([women.pop(0)]) for _ in range(num_teams)]
        # Add random players
        for i in range(num_teams):
            teams[i].add_players(pop_random_player(women, 0, len(women) - 1))

    # If there are baggages, add them to the teams and then balance the number of women and
    # total number of players
    if len(baggages) > 0:
        baggages.sort(reverse=True)
        mean_rank = sum(t.mean_rank for t in teams) / num_teams
        add_baggages_to_teams(teams, baggages, mean_rank)
        balance_teams(teams, men, women, mean_rank)

    # Add male players to the teams based on how team rankings compare to the average rank
    team_index = assign_players(mean_rank, men, teams, num_teams)

    # Add female players to the teams based on how team rankings compare to the average rank
    team_index = assign_players(mean_rank, women, teams, num_teams, team_index)

    # Convert PlayerGroups to Team objects
    final_teams = []
    for player_group in teams:
        final_teams.append(Team(player_group.players.copy()))

    # Export teams to Excel
    timestamp = dt.datetime.now().strftime('%m-%d-%Y_%H-%M-%S')
    save_path = os.path.join(save_directory, f'teams_{timestamp}.xlsx')

    wb = Workbook()
    ws = wb.active
    offset = 0
    for team in final_teams:
        team_data = [{'name': 'Name', 'gender': 'Gender', 'rank': 'Rank'}] + [p.to_dict() for p in team.players] + [{'name': 'Average', 'gender': '', 'rank': team.mean_rank}]
        for row_idx, row in enumerate(team_data, start=offset + 1):
            for col_idx, (key, value) in enumerate(row.items(), start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        offset += len(team_data) + 4
    wb.save(save_path)


def export_players(players: list[Player], save_directory: str):
    """Export player data to Excel"""
    timestamp = dt.datetime.now().strftime('%m-%d-%Y_%H-%M-%S')
    save_path = os.path.join(save_directory, f'players_{timestamp}.xlsx')
    wb = Workbook()
    ws = wb.active
    data = [{'name': 'Name', 'gender': 'Gender', 'rank': 'Rank'}] + [p.to_dict() for p in players]
    for row_idx, row in enumerate(data, start=1):
        for col_idx, (key, value) in enumerate(row.items(), start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    wb.save(save_path)


def load_exported_players(filepath: str) -> Roster:
    """Load a previously-exported players .xlsx file

    Normalizes the `name` column (stripping whitespace and title-casing) when present.
    """

    wb = load_workbook(filepath)
    ws = wb.active
    players = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
        if row[0]:  # Assuming name is first column
            name = str(row[0]).strip()
            gender = Gender(row[1]) if row[1] else Gender.MALE  # Default
            rank = int(row[2]) if row[2] else 0
            players.append(Player(name=name, gender=gender, rank=rank))
    return Roster(players)


def get_attendance_indices(roster: Roster, exported: Roster, key: str = 'name') -> tuple[list[int], list[str]]:
    """Return a list of indices in `roster` that match any name in `exported`.

    Matching is case- and whitespace-insensitive and uses the `name` column by default.
    Returns (indices, unmatched_exported_names).
    """
    if not roster or not exported:
        return [], []

    roster_names = {p.name.lower().replace(' ', ''): i for i, p in enumerate(roster.players)}
    matched_indices = []
    unmatched = []
    for p in exported.players:
        match_name = p.name.lower().replace(' ', '')
        if match_name in roster_names:
            matched_indices.append(roster_names[match_name])
        else:
            unmatched.append(p.name)
    return matched_indices, unmatched


def apply_attendance_column(roster: Roster, indices: list[int], column_name: str = 'attended') -> Roster:
    """Add or update a boolean attendance column on `roster` marking provided indices True"""
    for i in indices:
        if i < len(roster.players):
            setattr(roster.players[i], column_name, True)
    return roster
